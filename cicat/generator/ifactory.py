# -*- coding: utf-8 -*-
"""
::::::::::::::::::::::::  Critical Infrastructure Cyberspace Analysis Tool (CICAT)  :::::::::::::::::::::::::::::::::::::::

                                            NOTICE
                                            
The contents of this material reflect the views of the author and/or the Director of the Center for Advanced Aviation 
System Development (CAASD), and do not necessarily reflect the views of the Federal Aviation Administration (FAA) 
or the Department of Transportation (DOT). Neither the FAA nor the DOT makes any warranty or guarantee, or promise, 
expressed or implied, concerning the content or accuracy of the views expressed herein. 

This is the copyright work of The MITRE Corporation and was produced for the U.S. Government under Contract Number 
DTFAWA-10-C-00080 and is subject to Federal Aviation Administration Acquisition Management System Clause 3.5-13, 
Rights in Data-General, Alt. III and Alt. IV (Oct. 1996). No other use other than that granted to the U.S. Government, 
or to those acting on behalf of the U.S. Government, under that Clause is authorized without the express written permission 
of The MITRE Corporation. For further information, please contact The MITRE Corporation, Contract Office, 7515 Colshire Drive, 
McLean, VA 22102 (703) 983-6000. ©2020 The MITRE Corporation. 

The Government retains a nonexclusive, royalty-free right to publish or reproduce this document, or to allow others to do so, for 
“Government Purposes Only.”                                           
                                            
(c) 2020 The MITRE Corporation. All Rights Reserved.

::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
ifactory.py - factory class for infrastructure data
::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::::
"""


import openpyxl

from imodel import COMPONENT
from imodel import CTYPE
from imodel import SYSTEM
from imodel import FSMAP
from imodel import FX
from imodel import LOCATION
from imodel import CAPABILITY
from imodel import SURFACE
from imodel import CONNECTION
from columnloader import LoadColsInSpreadsheet

class CI_FACTORY():
    def __init__ (self, filename, trace ):
        self.filename = filename
        self.trace = trace
        if self.trace:
            print ('CI_FACTORY constructed..')
        
    def getLoader (self, sheetname ):

        if (sheetname == 'CTYPE'):
            return CTYPE_FACTORY (self.filename, sheetname, self.trace)
        elif (sheetname == 'COMPONENT'):
            return  COMPONENT_FACTORY(self.filename, sheetname, self.trace)
        elif (sheetname == 'CONNECTION'):
            return CONNECTION_FACTORY (self.filename, sheetname, self.trace )
        elif (sheetname == 'SYSTEM'):
            return SYSTEM_FACTORY(self.filename, sheetname, self.trace)
        elif (sheetname == 'FSMAP'):
            return FSMAP_FACTORY(self.filename, sheetname, self.trace)
        elif (sheetname == 'FUNCTION'):
            return FX_FACTORY (self.filename, sheetname, self.trace)
        elif (sheetname == 'LOCATION'):
            return LOCATION_FACTORY(self.filename, sheetname, self.trace )
        elif (sheetname == 'CAPABILITY'):
            return CAPABILITY_FACTORY(self.filename, sheetname, self.trace )
        elif (sheetname == 'SURFACE'):
            return SURFACE_FACTORY (self.filename, sheetname, self.trace )
        else:
            if self.trace:
                print ('WARNING! CI_FACTORY: unrecognized object:', sheetname )
       
    def initRelationships(self, myDATAWARE):
       if self.trace:
           print ('CI_FACTORY loading relationships..')
 
       if self.trace:
           print ('Link_2: Systems to Components'  )  
       for c in myDATAWARE['COMPONENT']:
           c.link_2 (myDATAWARE['SYSTEM'], self.trace)
        
       if self.trace:
           print ('Link_3: Components to Systems')  
       for r in myDATAWARE['SYSTEM']:
           r.link_3 (myDATAWARE['COMPONENT'], self.trace)    
           
       if self.trace:
           print ('Mapping Component Types to Controlled Interfaces')
       for r in myDATAWARE['CONNECTION']:
           r.mapCTYPE(myDATAWARE['CTYPE'])
                   
       if self.trace:
          print ('Link_4: Attack Surfaces to CTYPEs')
       for q in myDATAWARE['CTYPE']:
          q.link_4 (myDATAWARE['SURFACE'], self.trace)
               
       if self.trace:
          print ('Link_5: Functions to Systems using the FSMAP' )
       for r in myDATAWARE['SYSTEM']:
          r.link_5 (myDATAWARE['FSMAP'], myDATAWARE['FUNCTION'], self.trace )       
        
       if self.trace:
          print ('Link_6: Systems to Functions')
       for s in myDATAWARE['FUNCTION']:
          s.link_6 (myDATAWARE['SYSTEM'], self.trace )     
                
       if self.trace:
          print ('Link_7: Systems to Locations')
       for l in myDATAWARE['LOCATION']:
           l.link_7 (myDATAWARE['SYSTEM'], self.trace )      
        
       if self.trace:
          print ('Link_8: Capabilities to Functions') 
       for s in myDATAWARE['FUNCTION']:
           s.link_8 (myDATAWARE['CAPABILITY'], self.trace) 

       if self.trace:
          print ('Link_9: Functions to Capabilities') 
       for p in myDATAWARE['CAPABILITY']:
          p.link_9 (myDATAWARE['FUNCTION'], self.trace )               

       if self.trace:
          print ('CI_FACTORY relationships loaded.')      
    
class COMPONENT_FACTORY(CI_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('INDICATOR factory constructed..')
        return
    
    def findctype(self, ctID, ctArray):
        for j in ctArray:
            if ctID == j.getID():
                return j
            
    def load (self, ctypeList ):
          if self.trace:
              print ('Loading COMPONENT data..')
          book = openpyxl.load_workbook(self.filename, data_only=True) 
          sheet = book[self.sheetname]
          ret = []         
          for row in sheet.rows:
            ctype = self.findctype(row[1].value, ctypeList )           
            ret.append(COMPONENT (row[0].value, ctype, row[2].value, row[3].value, row[4].value, row[5].value  ))   
       
          del ret[0]
          return ret
 

class CTYPE_FACTORY(CI_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('CTYPE factory constructed..')
        return
    
    def load (self):
          if self.trace:
              print ('Loading CTYPE data..')
          book = openpyxl.load_workbook(self.filename, data_only=True) 
          sheet = book[self.sheetname]
          ret = []         
          for row in sheet.rows:
            ret.append(CTYPE (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value, row[5].value, row[6].value ))
            # CTYPEID, Vendor, Description, Type, Platform, Cpe, On
       
          del ret[0]
          return ret

class SYSTEM_FACTORY(CI_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('SYSTEM factory constructed..')
        return
    
    def load (self):
          if self.trace:
              print ('Loading SYSTEM data..')
          book = openpyxl.load_workbook(self.filename, data_only=True) 
          sheet = book[self.sheetname]
          ret = []         
          for row in sheet.rows:
            ret.append(SYSTEM (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value ))          

          del ret[0]
          return ret

class FSMAP_FACTORY(CI_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('FSMAP factory constructed..')
        return
    
    def load (self):
          if self.trace:
              print ('Loading FSMAP data..')
          book = openpyxl.load_workbook(self.filename, data_only=True) 
          sheet = book[self.sheetname]
          ret = []         
          for row in sheet.rows:
            ret.append(FSMAP (row[0].value, row[1].value  ))          

          del ret[0]
          return ret

class CONNECTION_FACTORY(CI_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('CONNECTION factory constructed..')
        return
    
    def load (self):
          if self.trace:
              print ('Loading CONNECTION data..')
          book = openpyxl.load_workbook(self.filename, data_only=True) 
          sheet = book[self.sheetname]
          ret = [] 
          cnt = 0
          for row in sheet.rows:
               entry = CONNECTION (cnt, row[0].value, row[1].value, row[2].value, 
                                             row[3].value, row[4].value, row[5].value  )            
               cnt = cnt+1
               ret.append (entry)
            
               # create second entry for reverse direction if not a oneway flow
               if not (entry.isOneway()):
                   entry = CONNECTION (cnt, row[1].value, row[0].value, row[2].value, 
                                             row[3].value, row[4].value, row[5].value  )
                   cnt = cnt+1
                   ret.append (entry)
                
          del ret[0]
          del ret[0]
          return ret
      

class LOCATION_FACTORY(CI_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('LOCATION factory constructed..')
        return
    
    def load (self):
          if self.trace:
              print ('Loading LOCATION data..')
          book = openpyxl.load_workbook(self.filename, data_only=True) 
          sheet = book[self.sheetname]
          ret = []         
          for row in sheet.rows:
            ret.append(LOCATION (row[0].value, row[1].value, row[2].value, row[3].value, row[4].value ))          

          del ret[0]
          return ret

class FX_FACTORY(CI_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('FX factory constructed..')
        return
    
    def load (self):
          if self.trace:
              print ('Loading FX data..')
          book = openpyxl.load_workbook(self.filename, data_only=True) 
          sheet = book[self.sheetname]
          ret = []         
          for row in sheet.rows:
            ret.append(FX (row[0].value, row[1].value, row[2].value, row[3].value ))          

          del ret[0]
          return ret

class CAPABILITY_FACTORY(CI_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('CAPABILITY factory constructed..')
        return
    
    def load (self):
          if self.trace:
              print ('Loading CAPABILITY data..')
          book = openpyxl.load_workbook(self.filename, data_only=True) 
          sheet = book[self.sheetname]
          ret = []         
          for row in sheet.rows:
            ret.append(CAPABILITY (row[0].value, row[1].value, row[2].value ))          

          del ret[0]
          return ret
      
class SURFACE_FACTORY(CI_FACTORY):
    def __init__ (self, filename, sheetname, trace):
        self.filename = filename
        self.sheetname = sheetname
        self.trace = trace
        if self.trace:
            print ('SURFACE factory constructed..')
        return
    
    def load (self):
        if self.trace:
           print ('Loading SURFACE data..')
                           
        ret = []
        surfacedict = LoadColsInSpreadsheet (self.filename, self.sheetname )
        for s in surfacedict.keys():
            if (s == 'MASTER') or (s == 'deny tag'):
                continue
            for l in surfacedict[s]:
               ret.append (SURFACE(s, l, 'Easy' ))
               
        return ret
    